import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time

chrome_options = Options()
chrome_options.binary_location = (r"C:\Program Files\Google\Chrome\Application\chrome.exe")  # Specify the path to the Chrome binary
chrome_options.add_argument("--headless")
chrome_options.add_argument("--lang=en-US")
chrome_options.add_argument("--disable-gpu")

service = Service(r'C:\Users\Dream\PycharmProjects\pythonProject\Demo\chromedriver.exe')  # Specify the path to your chromedriver
driver = webdriver.Chrome(service=service, options=chrome_options)

today=datetime.datetime.now()
day_name=today.strftime("%A")
print("Today is " + day_name)
# Open the Excel workbook
wb = load_workbook('excel_file.xlsx')  # Update with your Excel file path

# Initialize variables to store the longest and shortest options
longest_option = ""
shortest_option = None

def get_google_suggestions(keyword):
    # Open Google
    driver.get('https://www.google.com/?hl=en')
    # Find the search box
    search_box = driver.find_element('name', 'q')
    # Type the keyword in the search box
    search_box.send_keys(keyword)
    # Wait for suggestions to load
    time.sleep(2)  # Adjust as needed depending on your internet speed

    suggestions = driver.find_elements('css selector', 'li.sbct.PZPZlf')
    # Extract the text from each suggestion
    suggestion_texts = [suggestion.text.strip() for suggestion in suggestions if suggestion.text.strip()]

    return suggestion_texts

def find_longest_shortest(suggestions):
    if not suggestions:
        return None, None

    longest = max(suggestions, key=len)
    shortest = min(suggestions, key=len)
    return longest, shortest

sheet=wb[day_name]
print(f"Sheet name is : {sheet}")
for row in range(2, sheet.max_row + 1):  # Assuming data starts from the second row
    keyword = sheet.cell(row=row, column=1).value
    if keyword:
        get_google_suggestions(keyword)
        # Get suggestions
        suggestions = get_google_suggestions(keyword)

    # Find and print the longest and shortest suggestion
    longest, shortest = find_longest_shortest(suggestions)

    # Output the results
    print(keyword)
    print("Longest Suggestion:", longest if longest else "No valid suggestion found")
    print("Shortest Suggestion:", shortest if shortest else "No valid suggestion found")
    sheet.cell(row=row, column=2).value = longest
    sheet.cell(row=row, column=3).value = shortest

wb.save('excel_file.xlsx')
driver.quit()
